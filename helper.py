import boto3
import csv
import openpyxl
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from openpyxl.worksheet.table import Table as ExcelTable

class AWSResourceChecker:
    def __init__(self):
        self.s3 = boto3.client('s3')
        self.ec2 = boto3.client('ec2')
        self.rds = boto3.client('rds')
        self.iam = boto3.client('iam')
        self.csv_file = "aws_resource_report.csv"
        self.pdf_file = "aws_resource_report.pdf"
        self.xl_file = "aws_resource_report.xlsx"
        self.csv_writer = csv.writer(open(self.csv_file, 'w', newline=''))
        self.csv_file_path = 'aws_resource_report.csv'
        self.pdf_file_path = 'aws_resource_report.pdf'
        self.csv_writer.writerow(["Resource Type", "Resource Name", "Status"])

    def check_and_save_resources(self):
        self.check_and_save_s3_buckets()
        self.check_and_save_security_groups()
        self.check_and_save_rds_instances()
        self.check_cloudtrail_on_ec2_instances()
        self.csv_to_pdf(self.csv_file_path, self.pdf_file_path)

        # self.check_iam_users_and_permissions()


    def check_and_save_s3_buckets(self):
        response = self.s3.list_buckets()
        row_data = []

        for bucket in response['Buckets']:
            bucket_name = bucket['Name']
            public_access_block = self.get_public_access_block(bucket_name)

            if not public_access_block:
                try:
                    website_configuration = self.s3.get_bucket_website(Bucket=bucket_name)
                    print(f"The S3 bucket {bucket_name} is configured for static website hosting.")
                    # You can add more information about the website configuration if needed
                except self.s3.exceptions.NoSuchWebsiteConfiguration:
                    print(f"The S3 bucket {bucket_name} is not configured for static website hosting.")
                    row_data.append(bucket_name)
        if len(row_data) == 0:
            self.csv_writer.writerow(["S3", "Proper", "NO Public Access Enabled"])
        else:
            self.csv_writer.writerow(["S3", str(row_data), "Public Access Disabled"])

    def check_and_save_security_groups(self):
        response = self.ec2.describe_security_groups()
        row_data = []

        for security_group in response['SecurityGroups']:
            group_id = security_group['GroupId']
            group_name = security_group['GroupName']

            instances_attached = self.check_ec2_instances_attached(group_id)

            for ingress_rule in security_group.get('IpPermissions', []):
                for ip_range in ingress_rule.get('IpRanges', []):
                    if ip_range['CidrIp'] == '0.0.0.0/0' and self.check_ec2_instances_attached(group_id):
                        rule_description = f"{ingress_rule['IpProtocol']} from {ip_range['CidrIp']}"
                        print(f"Security Group {group_name} ({group_id}) allows {rule_description}.")
                        row_data.append(group_name)

        if len(row_data) == 0:
            self.csv_writer.writerow(["Security Group", "none", "No issues"])
        else:
            self.csv_writer.writerow(["Security Group", str(row_data), " Allowed 0.0.0.0 Access"])

    def check_ec2_instances_attached(self, group_id):
        ec2_instances = self.ec2.describe_instances(Filters=[{'Name': 'instance.group-id', 'Values': [group_id]}])

        if ec2_instances['Reservations']:
            instance_ids = [instance['InstanceId'] for reservation in ec2_instances['Reservations'] for instance in reservation['Instances']]
            return True
        else:
            return False

    def get_public_access_block(self, bucket_name):
        try:
            response = self.s3.get_public_access_block(Bucket=bucket_name)
            return response['PublicAccessBlockConfiguration']
        except self.s3.exceptions.NoSuchPublicAccessBlockConfiguration:
            return None

    def check_cloudtrail_on_ec2_instances(self):
        ec2 = boto3.client('ec2')
        response = ec2.describe_instances()
        row_data = []

        for reservation in response['Reservations']:
            for instance in reservation['Instances']:
                instance_id = instance['InstanceId']
                cloudtrail_enabled = any(tag['Key'] == 'cloudtrail' and tag['Value'].lower() == 'enabled' for tag in instance.get('Tags', []))

                if cloudtrail_enabled:
                    print(f"CloudTrail is enabled for EC2 instance {instance_id}.")
                else:
                    row_data.append(instance_id)

        if len(row_data) == 0:
            self.csv_writer.writerow(["Cloud trial", "none", "No issues"])
        else:
            self.csv_writer.writerow(['Cloud trial', str(row_data), "These are no issues"])

    def check_and_save_rds_instances(self):
        response = self.rds.describe_db_instances()
        row_data = []

        for db_instance in response['DBInstances']:
            db_instance_identifier = db_instance['DBInstanceIdentifier']
            read_replica_status = db_instance.get('ReadReplicaDBInstanceIdentifiers', [])

            if not read_replica_status:
                row_data.append(db_instance_identifier)

        if len(row_data) == 0:
            self.csv_writer.writerow(["RDS", "Proper", "Read replicas enabled for all instances"])
        else:
            self.csv_writer.writerow(["RDS", str(row_data), "Read replicas not enabled"])

    def check_iam_users_and_permissions(self):
        try:
            # List IAM users
            response = self.iam.list_users()

            for user in response['Users']:
                user_name = user['UserName']
                attached_policies = []
                inline_policies = []

                # List attached IAM policies
                attached_policies_response = self.iam.list_attached_user_policies(UserName=user_name)
                for attached_policy in attached_policies_response['AttachedPolicies']:
                    attached_policies.append(attached_policy['PolicyName'])

                # List IAM user inline policies
                inline_policies_response = self.iam.list_user_policies(UserName=user_name)
                for inline_policy_name in inline_policies_response['PolicyNames']:
                    inline_policies.append(inline_policy_name)

                # Write the information to the CSV file
                self.csv_writer.writerow([user_name, ', '.join(attached_policies), ', '.join(inline_policies)])

        except Exception as e:
            print(f"Error: {e}")
    def csv_to_pdf(self,csv_file_path, pdf_file_path):
    # Read data from CSV file using pandas
        csv_file_path = "./aws_resource_report2.csv"
        pdf_file_path = "./aws_resource_report2.pdf"
        try:
            df = pd.read_csv(csv_file_path)
            data = [df.columns.tolist()] + df.values.tolist()
        except pd.errors.EmptyDataError:
            print("CSV file is empty.")
            return
        except Exception as e:
            print(f"Error reading CSV file: {e}")
            return

        # Create a PDF document
        doc = SimpleDocTemplate(pdf_file_path, pagesize=letter)
        table = Table(data)

        # Apply table styling
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), 'grey'),
            ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), 'white'),
            ('GRID', (0, 0), (-1, -1), 0.5, 'black')
        ]))

        # Build the PDF document
        doc.build([table])
        print(f"PDF report generated: {pdf_file_path}")
    def convert_csv_to_excel(self):
        # Read data from CSV file
        with open(self.csv_file, 'r') as csv_file:
            data = [row for row in csv.reader(csv_file)]

        # Create an Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write data to Excel sheet
        for row in data:
            sheet.append(row)

        # Save the Excel file
        workbook.save(self.xl_file)
        print(f"Excel file generated: {self.xl_file}")

    def create_pdf(self, excel_file_path, pdf_file_path):
        # Read data from Excel file
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        data = [list(row) for row in sheet.iter_rows(values_only=True)]

        if not data:
            print("No data found in the Excel file.")
            return

        # Create a PDF document
        doc = SimpleDocTemplate(pdf_file_path, pagesize=letter)
        table = Table(data)

        # Apply table styling
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), 'grey'),
            ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), 'white'),
            ('GRID', (0, 0), (-1, -1), 0.5, 'black')
        ]))

        # Build the PDF document
        doc.build([table])
        print(f"PDF report generated: {pdf_file_path}")




# Create an instance of AWSResourceChecker and check/save resources
awss = AWSResourceChecker()
awss.check_and_save_resources()
